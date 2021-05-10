import requests
import os
from bs4 import BeautifulSoup
import re
import openpyxl

path = r'C:\Users\hon\Desktop\PythonScript\stock_new.xlsx'

# path = os.getcwd()
# print(path)
# html_raw = requests.get('https://www.jisilu.cn/web/data/cb/list')
html_raw = requests.get('http://www.ninwin.cn/index.php?m=cb&show_cb_only=Y&show_listed_only=Y', headers = {
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'})
# print(html_raw.text)
mysoup = BeautifulSoup(html_raw.content,'lxml')
trs = mysoup.find_all('tr')
stock_names = []
dept_codes = []
prices_zhai = []
prices_gu = []
yijias = []
years = []
remain_amounts = []
pbs = []
ShuiQianShouYiLvs = []
ShuiQianHuiShous = []
BackValues = []
PureDebtValues = []
elas = []
crds = []

stock_codes =[]
debt_names = []
sub_industrys = []
stock_prices = []
remain_benxis = []
days_befores = []
huishou_years = []
# print(trs[2].find_all('td', attrs = {'class': 'bond_code_id industry'})[1])

for tr in trs[2:]:
    # 4 股票名称
    stock_name = tr.find('td', attrs = {'class': 'stock_name_id stock_name'})
    stock_names.append(stock_name.string)
    # 1 转债代码
    debt_code = tr.find('td', attrs = {'class': 'bond_code_id bond_code'})
    dept_codes.append(debt_code.string)
    # 6 转债价格
    price_zhai = tr.find('td', attrs = {'class': 'cb_price2_id'})
    prices_zhai.append(price_zhai.string)
    # 9 转股价格
    price_gu = tr.find('td', attrs = {'class': 'cb_strike_id'})
    prices_gu.append(price_gu.string)
    # 10 转股溢价率
    yijia = tr.find('td', attrs = {'class': 'cb_premium_id'})
    yijias.append(yijia.string)
    # 12 剩余年限
    year = tr.find('td', attrs = {'class': 'cb_t_id bond_t bond_t1'})
    if(not year.find('font')): # 不存在天数
        years.append(year.string)
    else: #存在天数
        day = year.find('font')
        day_str = day.string
        year.font.decompose()
        yr_str = year.string
        if(yr_str is None):# 连年都没有，小于一年
            years.append(day_str)
        else:# 有年有天
            years.append(yr_str + day_str)
    # 14 转债余额
    remain_amount = tr.find('td', attrs = {'class': 'stock_price_id remain_amount'})
    remain_amounts.append(remain_amount.string)
    # 15 P/B
    pb = tr.find('td', attrs = {'class': 'cb_elasticity_id'})
    pbs.append(pb.string)
    # 税前收益率
    ShuiQianShouYiLv = tr.find('td', attrs = {'class': 'cb_BT_id BT_yield'})
    ShuiQianShouYiLvs.append(ShuiQianShouYiLv.string)
    # 税前回售收益
    ShuiQianHuiShou = tr.find('td', attrs = {'class': 'cb_AT_id BT_red'})
    ShuiQianHuiShous.append(ShuiQianHuiShou.string)
    # 回售价值
    BackValue = tr.find('td', attrs = {'class': 'cb_value_id npv_red'})
    BackValues.append(BackValue.string)
    # 16 纯债务价值
    PureDebtValue = tr.find('td', attrs = {'class': 'cb_value_id npv_value'})
    PureDebtValues.append(PureDebtValue.string)
    # 弹性
    ela = tr.find('td', attrs = {'class': 'cb_elasticity_id elasticity'})
    elas.append(ela.string)
    # 17 信用
    crd = tr.find('td', attrs = {'class': 'bond_rating_id rating'})
    crds.append(crd.string)
#new added
    # 3 股票代码
    stock_code = tr.find('td', attrs = {'class': "bond_code_id stock_code"})
    stock_codes.append(stock_code.string)
    # 2 转债名称
    debt_name = tr.find('a', attrs = {'target': "_blank"})
    debt_names.append(debt_name.text)# 用text而不是string否则会存在None type
    # 5 子行业
    sub_industry = tr.find_all('td', attrs = {'class': 'bond_code_id industry'})[1]
    sub_industrys.append(sub_industry.string)
    # 7 股价
    stock_price = tr.find('td', attrs = {'class': 'stock_price_id'})
    stock_prices.append(stock_price.string)
    # 8 剩余本息
    remain_benxi = tr.find_all('td', attrs = {'class': 'cb_price2_id'})[1]
    remain_benxis.append(remain_benxi.string)
    # 11 距离转股日
    days_before = tr.find('td', attrs = {'class': 'cb_t_id'})
    days_befores.append(days_before.string)
    # 13 回售年限
    huishou_year = tr.find('td', attrs = {'class': 'cb_t_id red_t'})
    huishou_years.append(huishou_year.text)


    print(huishou_year.text)


# for i in range(len(dept_codes)):
#     print(dept_codes[i] + "   " + stock_names[i] + "   " + prices_zhai[i] + "   " + prices_gu[i] + "   " + yijias[i] + "   " + years[i] + "   " + remain_amounts[i] +  "   " + pbs[i] +  "   "
#     + ShuiQianShouYiLvs[i] +  "   " + ShuiQianHuiShous[i] +  "   " + BackValues[i] +  "   " + PureDebtValues[i] +  "   " + elas[i] +  "   " + crds[i])


all_in = []
for i in range(len(dept_codes)):
    a_stock_all_info = []

    a_stock_all_info.append(dept_codes[i])
    a_stock_all_info.append(debt_names[i])
    a_stock_all_info.append(stock_codes[i])
    a_stock_all_info.append(stock_names[i])
    a_stock_all_info.append(sub_industrys[i])
    a_stock_all_info.append(prices_zhai[i])
    a_stock_all_info.append(stock_prices[i])
    a_stock_all_info.append(remain_benxis[i])
    a_stock_all_info.append(prices_gu[i])
    a_stock_all_info.append(yijias[i])
    a_stock_all_info.append(days_befores[i])
    a_stock_all_info.append(years[i])
    a_stock_all_info.append(huishou_years[i])
    a_stock_all_info.append(remain_amounts[i])
    a_stock_all_info.append(pbs[i])
    a_stock_all_info.append(PureDebtValues[i])
    a_stock_all_info.append(crds[i])


    # a_stock_all_info.append(ShuiQianShouYiLvs[i])
    # a_stock_all_info.append(ShuiQianHuiShous[i])
    # a_stock_all_info.append(BackValues[i])
    # a_stock_all_info.append(elas[i])

    
    all_in.append(a_stock_all_info)
    print(a_stock_all_info)







wb = openpyxl.load_workbook(path)
sheet = wb.worksheets[0]
sheet.cell(1,1).value='转债代码'
sheet.cell(1,2).value='转债名称'
sheet.cell(1,3).value='股票代码'
sheet.cell(1,4).value='股票名称'
sheet.cell(1,5).value='子行业'
sheet.cell(1,6).value='转债价格'
sheet.cell(1,7).value='股价'
sheet.cell(1,8).value='剩余本息'
sheet.cell(1,9).value='转股价格'
sheet.cell(1,10).value='转股溢价率'
sheet.cell(1,11).value='距离转股日'
sheet.cell(1,12).value='剩余年限'
sheet.cell(1,13).value='回售年限'
sheet.cell(1,14).value='转债余额'
sheet.cell(1,15).value='P/B'
sheet.cell(1,16).value='纯债务价值'
sheet.cell(1,17).value='信用'

# sheet.cell(1,9).value='税前收益率'
# sheet.cell(1,10).value='税前回售收益'
# sheet.cell(1,11).value='回售价值'
# sheet.cell(1,13).value='弹性'


for row in range(len(dept_codes) + 2)[2:]:
    for col in range(1,18):
        sheet.cell(row,col).value = all_in[row - 2][col - 1]



wb.save(path)