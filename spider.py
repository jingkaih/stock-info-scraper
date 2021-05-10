import requests
import os
from bs4 import BeautifulSoup
import re
import openpyxl

path = r'C:\Users\hon\Desktop\PythonScript\stock.xlsx'

# path = os.getcwd()
# print(path)
# html_raw = requests.get('https://www.jisilu.cn/web/data/cb/list')
html_raw = requests.get('http://www.ninwin.cn/index.php?m=cb&show_cb_only=Y&show_listed_only=Y', headers = {
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'})
# print(html_raw.text)
mysoup = BeautifulSoup(html_raw.content,'lxml')
trs = mysoup.find_all('tr')
names = []
codes = []
prices_zhai = []
prices_gu = []
yijias = []
years = []
remain_amounts = []
pbs = []
ShuiQianShouYiLvs = []
ShuiQianHuiShous = []
BackValues = []
PureDebtValues = []
elas = []
crds = []
for tr in trs[2:]:
    name = tr.find('td', attrs = {'class': 'stock_name_id stock_name'})
    names.append(name.string)
    code = tr.find('td', attrs = {'class': 'bond_code_id bond_code'})
    codes.append(code.string)
    price_zhai = tr.find('td', attrs = {'class': 'cb_price2_id'})
    prices_zhai.append(price_zhai.string)
    price_gu = tr.find('td', attrs = {'class': 'cb_strike_id'})
    prices_gu.append(price_gu.string)
    yijia = tr.find('td', attrs = {'class': 'cb_premium_id'})
    yijias.append(yijia.string)
    year = tr.find('td', attrs = {'class': 'cb_t_id bond_t bond_t1'})
    if(not year.find('font')): # 不存在天数
        years.append(year.string)
    else: #存在天数
        day = year.find('font')
        day_str = day.string
        year.font.decompose()
        yr_str = year.string
        if(yr_str is None):# 连年都没有，小于一年
            years.append(day_str)
        else:# 有年有天
            years.append(yr_str + day_str)
    remain_amount = tr.find('td', attrs = {'class': 'stock_price_id remain_amount'})
    remain_amounts.append(remain_amount.string)
    pb = tr.find('td', attrs = {'class': 'cb_elasticity_id'})
    pbs.append(pb.string)
    ShuiQianShouYiLv = tr.find('td', attrs = {'class': 'cb_BT_id BT_yield'})
    ShuiQianShouYiLvs.append(ShuiQianShouYiLv.string)
    ShuiQianHuiShou = tr.find('td', attrs = {'class': 'cb_AT_id BT_red'})
    ShuiQianHuiShous.append(ShuiQianHuiShou.string)
    BackValue = tr.find('td', attrs = {'class': 'cb_value_id npv_red'})
    BackValues.append(BackValue.string)
    PureDebtValue = tr.find('td', attrs = {'class': 'cb_value_id npv_value'})
    PureDebtValues.append(PureDebtValue.string)
    ela = tr.find('td', attrs = {'class': 'cb_elasticity_id elasticity'})
    elas.append(ela.string)
    crd = tr.find('td', attrs = {'class': 'bond_rating_id rating'})
    crds.append(crd.string)



# for i in range(len(codes)):
#     print(codes[i] + "   " + names[i] + "   " + prices_zhai[i] + "   " + prices_gu[i] + "   " + yijias[i] + "   " + years[i] + "   " + remain_amounts[i] +  "   " + pbs[i] +  "   "
#     + ShuiQianShouYiLvs[i] +  "   " + ShuiQianHuiShous[i] +  "   " + BackValues[i] +  "   " + PureDebtValues[i] +  "   " + elas[i] +  "   " + crds[i])

all_in = []
for i in range(len(codes)):
    a_stock_all_info = []
    a_stock_all_info.append(codes[i])
    a_stock_all_info.append(names[i])
    a_stock_all_info.append(prices_zhai[i])
    a_stock_all_info.append(prices_gu[i])
    a_stock_all_info.append(yijias[i])
    a_stock_all_info.append(years[i])
    a_stock_all_info.append(remain_amounts[i])
    a_stock_all_info.append(pbs[i])
    a_stock_all_info.append(ShuiQianShouYiLvs[i])
    a_stock_all_info.append(ShuiQianHuiShous[i])
    a_stock_all_info.append(BackValues[i])
    a_stock_all_info.append(PureDebtValues[i])
    a_stock_all_info.append(elas[i])
    a_stock_all_info.append(crds[i])
    all_in.append(a_stock_all_info)


wb = openpyxl.load_workbook(path)
sheet = wb.worksheets[0]
sheet.cell(1,1).value='转债代码'
sheet.cell(1,2).value='转债名称'
sheet.cell(1,3).value='转债价格'
sheet.cell(1,4).value='转股价格'
sheet.cell(1,5).value='转股溢价率'
sheet.cell(1,6).value='剩余年限'
sheet.cell(1,7).value='转债余额'
sheet.cell(1,8).value='P/B'
sheet.cell(1,9).value='税前收益率'
sheet.cell(1,10).value='税前回售收益'
sheet.cell(1,11).value='回售价值'
sheet.cell(1,12).value='纯债务价值'
sheet.cell(1,13).value='弹性'
sheet.cell(1,14).value='信用'



for row in range(len(codes) + 2)[2:]:
    for col in range(1,15):
        sheet.cell(row,col).value = all_in[row - 2][col - 1]



wb.save(path)